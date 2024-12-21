import time
from playwright.sync_api import sync_playwright
from openpyxl import load_workbook

# Флаг отладки
DEBUG_MODE = False  # Установите в True, чтобы включить отладочные сообщения


def debug_print(message):
    if DEBUG_MODE:
        print(f"DEBUG: {message}")


def extract_data_from_page(page):
    """
    Извлекает информацию о счёте всех геймов с текущей страницы (point-by-point),
    а также информацию о подающем игроке в первом гейме первого сета.
    """
    game_data = []
    server_info = "Неизвестно"

    try:
        for i in range(2, 27, 2):
            score_selector = f"#detail > div.matchHistoryRowWrapper > div:nth-child({i}) > div.matchHistoryRow__scoreBox"
            try:
                score_element = page.locator(score_selector)
                if score_element.is_visible():
                    game_data.append(score_element.inner_text().replace("\n", ""))

                    if i == 2:
                        serve_selector_left = f"#detail > div.matchHistoryRowWrapper > div:nth-child({i}) > div.matchHistoryRow__servis.matchHistoryRow__home > div > svg"
                        serve_selector_right = f"#detail > div.matchHistoryRowWrapper > div:nth-child({i}) > div.matchHistoryRow__servis.matchHistoryRow__away > div > svg"

                        if page.locator(serve_selector_left).is_visible():
                            server_info = "Игрок 1 подает"
                        elif page.locator(serve_selector_right).is_visible():
                            server_info = "Игрок 2 подает"
            except Exception:
                break

        return game_data, server_info
    except Exception as e:
        print(f"Ошибка при извлечении данных: {e}")
        return [], "Неизвестно"


def convert_score_to_letters(scores):
    """
    Преобразует список счётов в формате ['X-Y', ...] в строку букв,
    определяя прирост выигрышей (A или B).
    """
    letters = []
    previous_score = [0, 0]

    for score in scores:
        current_score = list(map(int, score.split('-')))
        delta_player1 = current_score[0] - previous_score[0]
        delta_player2 = current_score[1] - previous_score[1]

        if delta_player1 > delta_player2:
            letters.append('A')
        elif delta_player2 > delta_player1:
            letters.append('B')

        previous_score = current_score

    return ''.join(letters)


def switch_tabs_and_collect_data(page):
    """
    Последовательно переключается между вкладками и собирает информацию с каждой.
    """
    all_data = {}
    try:
        tab_buttons = page.locator("#detail > div.subFilterOver.subFilterOver--indent > div > a > button")
        num_tabs = tab_buttons.count()

        for i in range(num_tabs):
            try:
                tab_buttons.nth(i).click()
                page.wait_for_selector("#detail > div.matchHistoryRowWrapper", timeout=2000)

                data, server_info = extract_data_from_page(page)
                data_letters = convert_score_to_letters(data)

                all_data[f"point-by-point/{i}"] = data_letters
                if i == 0:
                    all_data["server_info"] = server_info
            except Exception as e:
                print(f"Ошибка при переходе на вкладку {i}: {e}")
    except Exception as e:
        print(f"Ошибка при работе с вкладками: {e}")

    return all_data


def write_to_excel(file_path, parsed_data, match_url):
    """
    Записывает собранные данные в Excel-файл.
    """
    try:
        workbook = load_workbook(file_path)
        sheet = workbook["Sets"]

        # Ищем последнюю строку с данными в колонке 12
        last_row = sheet.max_row
        while sheet.cell(row=last_row, column=12).value is None and last_row > 1:
            last_row -= 1

        server_info = parsed_data.get("server_info", "Неизвестно")
        server_value = "A" if server_info == "Игрок 1 подает" else "B" if server_info == "Игрок 2 подает" else ""
        target_row = last_row + 1

        for set_index, (key, score) in enumerate(parsed_data.items()):
            sheet.cell(row=target_row, column=4).value = match_url
            if key == "server_info":
                continue

            if set_index == 0 and server_value:
                sheet.cell(row=target_row, column=9).value = server_value

            for i, letter in enumerate(score, start=12):
                sheet.cell(row=target_row, column=i).value = letter
            target_row = target_row + 1

        workbook.save(file_path)

    except Exception as e:
        print(f"Ошибка при записи в файл: {e}")


def process_all_match_links(excel_file_path):
    """
    Читает ссылки на матчи с листа MatchLinks и обрабатывает их.
    """
    workbook = load_workbook(excel_file_path)
    sheet = workbook["MatchLinks"]

    # Получаем все ссылки из первого столбца
    match_links = [sheet.cell(row=i, column=1).value for i in range(2, sheet.max_row + 1) if
                   sheet.cell(row=i, column=1).value]

    for match_url in match_links:
        process_match_page(match_url, excel_file_path)


def process_match_page(match_url, excel_file_path):
    """
    Переходит на страницу матча, собирает данные со всех вкладок point-by-point и записывает в Excel.
    """
    print(f"Обработка матча: {match_url}")

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        context = browser.new_context()
        page = context.new_page()

        try:
            page.goto(match_url)
            page.wait_for_selector("#detail > div.filterOver.filterOver--indent > div > a:nth-child(3) > button", timeout=5000)

            try:
                button_selector = "#detail > div.filterOver.filterOver--indent > div > a:nth-child(3) > button"
                page.locator(button_selector).click()
                page.wait_for_selector("#detail > div.matchHistoryRowWrapper", timeout=2000)
            except Exception as e:
                return

            all_data = switch_tabs_and_collect_data(page)

            write_to_excel(excel_file_path, all_data, match_url)

        except Exception as e:
            print(f"Ошибка при обработке матча {match_url}: {e}")
        finally:
            browser.close()


if __name__ == "__main__":
    start_time = time.time()
    excel_file_path = "C:\\Users\\User\\Desktop\\Python\\Projects\\7thGamePlayWright\\7thGamePW.xlsx"
    process_all_match_links(excel_file_path)
    total_time = time.time() - start_time
    print(f"Полное время выполнения программы: {total_time:.2f} секунд")
