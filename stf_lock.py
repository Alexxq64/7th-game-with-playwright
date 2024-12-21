import time
import asyncio
from asyncio import Semaphore, Lock
from playwright.async_api import async_playwright
from openpyxl import load_workbook
import random

DEBUG_MODE = False  # Установите True для включения отладочной информации
MAX_CONCURRENT_TASKS = 10  # Максимальное количество одновременно выполняемых задач
MAX_RETRIES = 3  # Количество повторных попыток при ошибках

def debug_print(message):
    if DEBUG_MODE:
        print(f"DEBUG: {message}")

async def extract_data_from_page(page, lock: Lock):
    """
    Извлекает данные из текущей вкладки
    """
    game_data = []
    server_info = "Неизвестно"

    async with lock:
        try:
            for i in range(2, 27, 2):  # Считаем строки с данными
                score_selector = f"#detail > div.matchHistoryRowWrapper > div:nth-child({i}) > div.matchHistoryRow__scoreBox"
                try:
                    score_element = page.locator(score_selector)
                    if await score_element.is_visible():
                        score_text = await score_element.inner_text()
                        game_data.append(score_text)

                        if i == 2:  # Определяем подающего в первом гейме
                            serve_selector_left = f"#detail > div.matchHistoryRowWrapper > div:nth-child({i}) > div.matchHistoryRow__servis.matchHistoryRow__home > div > svg"
                            serve_selector_right = f"#detail > div.matchHistoryRowWrapper > div:nth-child({i}) > div.matchHistoryRow__servis.matchHistoryRow__away > div > svg"

                            if await page.locator(serve_selector_left).is_visible():
                                server_info = "Игрок 1 подает"
                            elif await page.locator(serve_selector_right).is_visible():
                                server_info = "Игрок 2 подает"
                            debug_print(f"Информация о подающем игроке: {server_info}")
                except Exception as e:
                    debug_print(f"Ошибка при извлечении данных из строки {i}: {e}")
                    break

            return game_data, server_info
        except Exception as e:
            debug_print(f"Ошибка при извлечении данных: {e}")
            return [], "Неизвестно"

def convert_score_to_letters(scores):
    """
    Конвертирует счета в формат букв (A/B)
    """
    letters = []
    previous_score = [0, 0]

    for score in scores:
        try:
            current_score = list(map(int, score.split('-')))
            delta_player1 = current_score[0] - previous_score[0]
            delta_player2 = current_score[1] - previous_score[1]

            if delta_player1 > delta_player2:
                letters.append('A')
            elif delta_player2 > delta_player1:
                letters.append('B')

            previous_score = current_score
        except Exception as e:
            debug_print(f"Ошибка при конвертации счета '{score}': {e}")

    return ''.join(letters)

async def switch_tabs_and_collect_data(page, lock: Lock):
    """
    Переключается между вкладками и собирает данные
    """
    all_data = {}
    try:
        tab_buttons = page.locator("#detail > div.subFilterOver.subFilterOver--indent > div > a > button")
        num_tabs = await tab_buttons.count()
        debug_print(f"Найдено вкладок: {num_tabs}")

        for i in range(num_tabs):
            try:
                debug_print(f"Переход на вкладку {i}...")
                await tab_buttons.nth(i).click()
                await page.wait_for_selector("#detail > div.matchHistoryRowWrapper", timeout=5000)  # Увеличена задержка

                data, server_info = await extract_data_from_page(page, lock)
                data_letters = convert_score_to_letters(data)

                if not data_letters:
                    debug_print(f"Вкладка {i} пуста. Данные отсутствуют.")

                all_data[f"point-by-point/{i}"] = data_letters
                if i == 0:
                    all_data["server_info"] = server_info
            except Exception as e:
                debug_print(f"Ошибка при переходе на вкладку {i}: {e}")
    except Exception as e:
        debug_print(f"Ошибка при работе с вкладками: {e}")

    return all_data

async def retry_action(action, retries: int = MAX_RETRIES, delay: int = 1):
    """
    Повторяет действие несколько раз в случае ошибок
    """
    last_exception = None
    for attempt in range(retries):
        try:
            return await action()  # Пытаемся выполнить действие
        except Exception as e:
            last_exception = e
            debug_print(f"Ошибка при попытке {attempt + 1}: {e}")
            await asyncio.sleep(random.uniform(1, 3))  # Задержка перед повтором
    # После всех попыток выбрасываем исключение
    raise last_exception

def write_to_excel(file_path, parsed_data, match_url):
    """
    Записывает данные в Excel-файл
    """
    try:
        workbook = load_workbook(file_path)
        sheet = workbook["Sets"]

        last_row = sheet.max_row
        while sheet.cell(row=last_row, column=12).value is None and last_row > 1:
            last_row -= 1

        server_info = parsed_data.get("server_info", "Неизвестно")
        server_value = "A" if server_info == "Игрок 1 подает" else "B" if server_info == "Игрок 2 подает" else ""
        target_row = last_row + 1

        for set_index, (key, score) in enumerate(parsed_data.items()):
            if key == "server_info":
                continue

            sheet.cell(row=target_row, column=4).value = match_url

            if set_index == 0 and server_value:
                sheet.cell(row=target_row, column=9).value = server_value

            for i, letter in enumerate(score, start=12):
                sheet.cell(row=target_row, column=i).value = letter
            target_row += 1

        workbook.save(file_path)
        debug_print(f"Данные записаны в файл {file_path}.")
    except Exception as e:
        debug_print(f"Ошибка при записи в файл: {e}")

async def process_match_page(match_url, excel_file_path, browser, semaphore, lock):
    """
    Обрабатывает одну страницу матча с повторными попытками
    """
    async with semaphore:
        print(f"Обработка матча: {match_url}")
        page = await browser.new_page()

        try:
            await page.goto(match_url)
            await page.wait_for_selector("#detail > div.filterOver.filterOver--indent > div > a:nth-child(3) > button", timeout=5000)

            try:
                button_selector = "#detail > div.filterOver.filterOver--indent > div > a:nth-child(3) > button"
                await page.locator(button_selector).click()
                await page.wait_for_selector("#detail > div.matchHistoryRowWrapper", timeout=5000)  # Увеличена задержка
            except Exception as e:
                debug_print(f"Ошибка при нажатии кнопки point-by-point: {e}")
                return

            # Собираем данные с повторными попытками
            all_data = await retry_action(lambda: switch_tabs_and_collect_data(page, lock))

            if not all_data:
                debug_print(f"Для матча {match_url} данные не собраны.")
                return  # Прерываем обработку, если данные не собраны

            write_to_excel(excel_file_path, all_data, match_url)
        except Exception as e:
            debug_print(f"Ошибка при обработке матча {match_url}: {e}")
        finally:
            await page.close()

async def process_all_match_links(excel_file_path):
    """
    Обрабатывает все ссылки из файла Excel
    """
    workbook = load_workbook(excel_file_path)
    sheet = workbook["MatchLinks"]
    match_links = [sheet.cell(row=i, column=1).value for i in range(2, sheet.max_row + 1) if
                   sheet.cell(row=i, column=1).value]

    semaphore = Semaphore(MAX_CONCURRENT_TASKS)
    lock = Lock()

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)

        tasks = [process_match_page(url, excel_file_path, browser, semaphore, lock) for url in match_links]
        await asyncio.gather(*tasks)

        await browser.close()

if __name__ == "__main__":
    start_time = time.time()
    excel_file_path = "C:\\Users\\User\\Desktop\\Python\\Projects\\7thGamePlayWright\\7thGamePW.xlsx"
    asyncio.run(process_all_match_links(excel_file_path))
    print(f"Полное время выполнения программы: {time.time() - start_time:.2f} секунд")
