import asyncio
from playwright.async_api import async_playwright  # Используем асинхронный API Playwright
import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


async def save_match_links_playwright(tournament_url, output_to_file=True):
    """
    Извлекает ссылки на матчи с использованием Playwright и сохраняет их в Excel файл или выводит в консоль.

    Параметры:
        tournament_url (str): URL страницы турнира.
        output_to_file (bool): Если True, сохраняет в файл; если False, выводит в консоль.
    """
    # Запускаем Playwright
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)  # Открытие браузера в фоновом режиме
        page = await browser.new_page()

        try:
            # Открываем страницу турнира
            await page.goto(tournament_url)
            await page.wait_for_selector("a[href*='game-summary']", timeout=5000)  # Ожидание появления ссылок

            # Находим все элементы ссылок на матчи
            match_links = []
            elements = await page.query_selector_all("a[href*='game-summary']")

            for element in elements:
                match_url = await element.get_attribute("href")
                if match_url and match_url not in match_links:  # Уникальные ссылки
                    match_links.append(match_url)

            if output_to_file:
                # Путь к файлу Excel
                file_path = "C:\\Users\\User\\Desktop\\Python\\Projects\\7thGamePlayWright\\7thGamePW.xlsx"

                try:
                    # Загружаем файл Excel или создаем новый
                    if os.path.exists(file_path):
                        wb = load_workbook(file_path)
                    else:
                        wb = Workbook()

                    # Создаем новый лист, если его нет
                    if "MatchLinks" not in wb.sheetnames:
                        ws = wb.create_sheet("MatchLinks")
                        ws.cell(1, 1, "Ссылка на матч")  # Заголовок
                    else:
                        ws = wb["MatchLinks"]

                    # Начинаем запись с первой свободной строки
                    start_row = ws.max_row + 1
                    for link in match_links:
                        ws.cell(start_row, 1, link)
                        ws.cell(start_row, 1).hyperlink = link
                        start_row += 1

                    # Автоподгонка ширины столбцов
                    column = get_column_letter(1)
                    max_length = max(len(str(ws[f"{column}{r}"].value or "")) for r in range(1, ws.max_row + 1))
                    ws.column_dimensions[column].width = max_length + 2

                    # Сохраняем файл Excel
                    wb.save(file_path)
                    print(f"Ссылки на матчи сохранены в Excel файл по пути: {file_path}")

                except PermissionError:
                    print(f"Ошибка! Файл '{file_path}' в данный момент открыт. Закройте файл и попробуйте снова.")
                except Exception as e:
                    print(f"Произошла ошибка: {e}")

            else:
                # Выводим в консоль
                print("Ссылки на матчи:")
                for link in match_links:
                    print(link)

        finally:
            # Закрываем браузер
            await browser.close()


# Запуск асинхронного кода
if __name__ == "__main__":
    tournament_url = "https://www.livesport.com/tennis/atp-singles/antwerp/results/"
    # Выберите, куда выводить результаты: True для файла, False для консоли
    asyncio.run(save_match_links_playwright(tournament_url, output_to_file=True))
